#!/usr/bin/env python3
"""Apply a supported translation table to a game work tree and build patch.xp3."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import pathlib
import shutil
import subprocess
import sys
import time
from typing import Any

from build_patch import (
    EXPECTED_COLUMNS,
    create_xp3,
    detect_modified_resources,
    convert_modified_tlg_previews,
    merge_patch_entries,
    verify_patch,
)
from render_name_localization import (
    CANVAS_SIZE,
    make_contact_sheet,
    render_name,
)


NAME_REQUIRED_COLUMNS = {
    "file_name",
    "ini_key",
    "image_width_px",
    "image_height_px",
    "translation_zh_cn",
}


def sha256(path: pathlib.Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def unique_backup_dir(root: pathlib.Path, label: str) -> pathlib.Path:
    stamp = time.strftime("%Y%m%d_%H%M%S")
    candidate = root / "backups" / f"{label}_{stamp}"
    suffix = 2
    while candidate.exists():
        candidate = root / "backups" / f"{label}_{stamp}_{suffix}"
        suffix += 1
    candidate.mkdir(parents=True)
    return candidate


def reset_work_dir(work_root: pathlib.Path) -> pathlib.Path:
    work_dir = (work_root / "build" / "table_patch_work").resolve()
    expected_parent = (work_root / "build").resolve()
    if work_dir.parent != expected_parent or work_root not in work_dir.parents:
        raise ValueError("临时构建目录校验失败")
    if work_dir.exists():
        if work_dir.is_symlink():
            raise ValueError(f"拒绝删除符号链接形式的临时目录: {work_dir}")
        shutil.rmtree(work_dir)
    work_dir.mkdir(parents=True)
    return work_dir


def open_workbook(path: pathlib.Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl；请使用项目自带运行环境或安装 openpyxl") from exc
    return load_workbook(path, read_only=True, data_only=False)


def find_header_row(workbook, required: set[str]) -> tuple[Any, int, list[str]] | None:
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
            1,
        ):
            headers = [clean_cell(value) for value in row]
            if required.issubset(set(headers)):
                return sheet, row_number, headers
    return None


def detect_table(table_path: pathlib.Path) -> tuple[str, Any | None, tuple[Any, int, list[str]] | None]:
    suffix = table_path.suffix.lower()
    if suffix == ".tsv":
        return "scenario_tsv", None, None
    if suffix != ".xlsx":
        raise ValueError("只支持 .xlsx Excel 表格和 .tsv 剧情表格")

    workbook = open_workbook(table_path)
    scenario = find_header_row(workbook, set(EXPECTED_COLUMNS))
    if scenario is not None:
        return "scenario_xlsx", workbook, scenario
    name_images = find_header_row(workbook, NAME_REQUIRED_COLUMNS)
    if name_images is not None:
        return "name_image_xlsx", workbook, name_images
    workbook.close()
    raise ValueError(
        "无法识别此表格。目前支持：剧情翻译 Excel/TSV，以及包含 "
        "file_name、ini_key、image_width_px、image_height_px、translation_zh_cn "
        "的 image/name 姓名图片 Excel。"
    )


def validate_work_root(work_root: pathlib.Path, table_kind: str) -> None:
    manifest = work_root / "extracted" / "_xp3_manifest.csv"
    if not manifest.is_file():
        raise FileNotFoundError(
            "所选文件夹不是完整的游戏工作目录，缺少: "
            f"{manifest}"
        )
    if table_kind.startswith("scenario"):
        scenario_dir = work_root / "extracted" / "_scenario_decompiled"
        if not scenario_dir.is_dir():
            raise FileNotFoundError(f"剧情反编译目录不存在: {scenario_dir}")
        text_root = work_root / "localization" / "text_utf8"
        if not text_root.is_dir():
            raise FileNotFoundError(f"UTF-8 文本目录不存在: {text_root}")


def choose_font(explicit_font: pathlib.Path | None) -> pathlib.Path:
    candidates = []
    if explicit_font is not None:
        candidates.append(explicit_font)
    candidates.extend(
        [
            pathlib.Path(r"C:\Windows\Fonts\NotoSansSC-VF.ttf"),
            pathlib.Path(r"C:\Windows\Fonts\msyhbd.ttc"),
            pathlib.Path(r"C:\Windows\Fonts\msyh.ttc"),
        ]
    )
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()
    raise FileNotFoundError("未找到可用的简体中文字体（Noto Sans SC 或微软雅黑）")


def rows_from_header(sheet, header_row: int, headers: list[str]) -> list[dict[str, Any]]:
    indexes = {header: index for index, header in enumerate(headers) if header}
    results: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {
            header: values[index] if index < len(values) else None
            for header, index in indexes.items()
        }
        if any(clean_cell(value) for value in row.values()):
            results.append(row)
    return results


def apply_name_table(
    workbook,
    header_info: tuple[Any, int, list[str]],
    work_root: pathlib.Path,
    work_dir: pathlib.Path,
    font_path: pathlib.Path,
) -> dict[str, Any]:
    sheet, header_row, headers = header_info
    rows = rows_from_header(sheet, header_row, headers)
    entries = []
    for row_number, row in enumerate(rows, header_row + 1):
        file_name = clean_cell(row.get("file_name"))
        translation = clean_cell(row.get("translation_zh_cn"))
        if not file_name or not translation:
            continue
        path_name = pathlib.PurePath(file_name)
        if path_name.name != file_name or pathlib.Path(file_name).suffix.lower() != ".png":
            raise ValueError(f"第 {row_number} 行的 file_name 非法: {file_name!r}")
        entries.append(
            {
                "file_name": file_name,
                "original_text": clean_cell(row.get("original_text")),
                "translation_zh_cn": translation,
                "translator_note": clean_cell(row.get("translator_note")),
            }
        )
    if not entries:
        raise ValueError("表格中没有填写 translation_zh_cn 的姓名图片行")
    file_names = [entry["file_name"].casefold() for entry in entries]
    if len(file_names) != len(set(file_names)):
        raise ValueError("表格中存在重复的 file_name")

    source_dir = work_root / "extracted" / "image" / "name"
    if not source_dir.is_dir():
        raise FileNotFoundError(f"姓名图片目录不存在: {source_dir}")
    render_dir = work_dir / "name_rendered"
    render_dir.mkdir(parents=True)
    rendered_records = []
    for entry in entries:
        source = source_dir / entry["file_name"]
        if not source.is_file():
            raise FileNotFoundError(f"表格对应的原图片不存在: {source}")
        from PIL import Image

        with Image.open(source) as image:
            dimensions = image.size
        if dimensions != CANVAS_SIZE:
            raise ValueError(f"图片尺寸不符合 406x66: {source} ({dimensions})")
        rendered, metrics = render_name(entry["translation_zh_cn"], font_path)
        destination = render_dir / entry["file_name"]
        rendered.save(destination, format="PNG", optimize=True)
        rendered_records.append({**entry, **metrics, "sha256": sha256(destination)})

    comparison = work_dir / "name_render_comparison.png"
    make_contact_sheet(entries, source_dir, render_dir, comparison)

    backup_dir = unique_backup_dir(work_root, "image_name_before_table_apply")
    changed = 0
    for entry in entries:
        source = source_dir / entry["file_name"]
        rendered = render_dir / entry["file_name"]
        shutil.copy2(source, backup_dir / entry["file_name"])
        if sha256(source) != sha256(rendered):
            changed += 1
        shutil.copy2(rendered, source)

    manifest = {
        "table_type": "name_image_xlsx",
        "translated_rows": len(entries),
        "changed_on_this_run": changed,
        "font": str(font_path),
        "target_directory": str(source_dir),
        "backup_directory": str(backup_dir),
        "comparison_image": str(comparison),
        "entries": rendered_records,
    }
    (work_dir / "name_apply_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def export_scenario_xlsx(
    header_info: tuple[Any, int, list[str]], output_path: pathlib.Path
) -> pathlib.Path:
    sheet, header_row, headers = header_info
    rows = rows_from_header(sheet, header_row, headers)
    if not rows:
        raise ValueError("剧情 Excel 中没有数据行")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=EXPECTED_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: clean_cell(row.get(column)) for column in EXPECTED_COLUMNS})
    return output_path


def backup_existing_patch(work_root: pathlib.Path) -> pathlib.Path | None:
    patch_path = work_root / "patch.xp3"
    if not patch_path.is_file():
        return None
    backup_dir = unique_backup_dir(work_root, "patch_before_table_build")
    backup = backup_dir / "patch.xp3"
    shutil.copy2(patch_path, backup)
    return backup


def build_resource_patch(
    work_root: pathlib.Path, tools_root: pathlib.Path, work_dir: pathlib.Path
) -> dict[str, Any]:
    print("[3/5] 检测所有相对原版发生变化的资源")
    resources = detect_modified_resources(work_root)
    tlg_resources: dict[str, pathlib.Path] = {}
    if (work_root / "extracted" / "_tlg_png").is_dir():
        print("[4/5] 转换发生变化的 TLG 预览")
        tlg_resources = convert_modified_tlg_previews(work_root, work_dir, tools_root)
    else:
        print("[4/5] 工作目录没有 TLG 预览，跳过 TLG 转换")
    entries = merge_patch_entries(
        [("修改资源", resources), ("修改 TLG", tlg_resources)]
    )
    if not entries:
        raise ValueError("未检测到任何相对原版发生变化的资源，无法生成补丁")

    print("[5/5] 生成并重新解包验证 patch.xp3")
    patch_path = work_root / "patch.xp3"
    report_entries = create_xp3(entries, patch_path)
    verify_patch(
        pathlib.Path(sys.executable),
        work_root,
        patch_path,
        work_dir,
        len(entries),
        tools_root,
    )
    return {
        "patch_path": str(patch_path),
        "patch_size": patch_path.stat().st_size,
        "patch_entries": len(entries),
        "modified_resources": len(resources),
        "modified_tlg": len(tlg_resources),
        "entries": report_entries,
    }


def run_scenario_builder(
    work_root: pathlib.Path,
    tools_root: pathlib.Path,
    tsv_path: pathlib.Path,
) -> dict[str, Any]:
    builder = tools_root / "build_patch.py"
    command = [
        sys.executable,
        str(builder),
        "--project-root",
        str(work_root),
        "--tsv",
        str(tsv_path),
        "--tools-root",
        str(tools_root),
        "--skip-utf8-system-text",
        "--gbk-help-text",
    ]
    subprocess.run(command, cwd=work_root, check=True)
    patch_path = work_root / "patch.xp3"
    report_path = work_root / "build" / "patch_build_report.json"
    return {
        "patch_path": str(patch_path),
        "patch_size": patch_path.stat().st_size,
        "scenario_report": str(report_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从翻译表格应用资源并为选定的游戏工作目录生成 patch.xp3"
    )
    parser.add_argument("--table", required=True, type=pathlib.Path)
    parser.add_argument("--work-root", required=True, type=pathlib.Path)
    parser.add_argument("--tools-root", required=True, type=pathlib.Path)
    parser.add_argument("--font", type=pathlib.Path)
    args = parser.parse_args()

    table_path = args.table.resolve()
    work_root = args.work_root.resolve()
    tools_root = args.tools_root.resolve()
    if not table_path.is_file():
        raise FileNotFoundError(f"表格不存在: {table_path}")
    if not work_root.is_dir():
        raise FileNotFoundError(f"游戏工作目录不存在: {work_root}")
    if not (tools_root / "build_patch.py").is_file():
        raise FileNotFoundError(f"工具目录不完整: {tools_root}")

    print("[1/5] 识别所选表格")
    table_kind, workbook, header_info = detect_table(table_path)
    print(f"  表格类型: {table_kind}")
    validate_work_root(work_root, table_kind)
    work_dir = reset_work_dir(work_root)
    old_patch_backup = backup_existing_patch(work_root)
    apply_report: dict[str, Any] | None = None

    try:
        if table_kind == "name_image_xlsx":
            assert workbook is not None and header_info is not None
            print("[2/5] 生成姓名图片、备份并替换工作目录资源")
            apply_report = apply_name_table(
                workbook,
                header_info,
                work_root,
                work_dir,
                choose_font(args.font),
            )
            build_report = build_resource_patch(work_root, tools_root, work_dir)
        else:
            print("[2/5] 准备剧情 TSV")
            if table_kind == "scenario_xlsx":
                assert workbook is not None and header_info is not None
                tsv_path = export_scenario_xlsx(
                    header_info, work_dir / "scenario_from_selected_excel.tsv"
                )
            else:
                tsv_path = table_path
            build_report = run_scenario_builder(work_root, tools_root, tsv_path)
    finally:
        if workbook is not None:
            workbook.close()

    report = {
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "selected_table": str(table_path),
        "table_type": table_kind,
        "work_root": str(work_root),
        "tools_root": str(tools_root),
        "previous_patch_backup": str(old_patch_backup) if old_patch_backup else None,
        "apply": apply_report,
        "build": build_report,
    }
    report_path = work_root / "build" / "table_patch_build_report.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print("\n表格应用和补丁构建成功！")
    print(f"补丁: {work_root / 'patch.xp3'}")
    print(f"报告: {report_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\n构建失败: {exc}", file=sys.stderr)
        raise
